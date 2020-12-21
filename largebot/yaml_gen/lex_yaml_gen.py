import argparse
import os
from shutil import copyfile
import sys
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from typing import Dict, List, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.worksheet.worksheet import Worksheet
from ruamel.yaml import YAML
from largebot.logger import get_logger


logger = get_logger(__name__)


__title__ = 'LexYamlGen'
__description__ = 'Converts Amazon Lex Bot Definition and Bot Template files to Yaml'
__version__ = '0.1.0'

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USER_DOCS_DIR = os.path.join(os.path.expanduser('~'), 'Documents')

# =========== Exceptions ===========

class Error(Exception):
    pass


class MissingValue(Error):
    def __init__(self, worksheet: Worksheet, cell: Cell) -> None:
        cell_id = f"{cell.column_letter}{cell.row}"
        super().__init__(
            f"Missing value in Worksheet '{worksheet.title}', cell '{cell_id}'.")


class IncompleteWorksheet(Error):
    def __init__(self, worksheet: Worksheet) -> None:
        super().__init__(f"The worksheet '{worksheet.title}' is incomplete.")


class InvalidIntentName(Error):
    def __init__(self, intent_name: str) -> None:
        super().__init__(
            f"The Intent Name '{intent_name}' is referenced but does not exist in the 'Intents' worksheet.")


class InvalidInputFile(Error):
    def __init__(self) -> None:
        super().__init__(f"The input file provided must be a valid Excel file.")

class IncompleteWorkbook(Error):
    def __init__(self) -> None:
        super().__init__(f"The Excel provided is missing required worksheets.")

# ==================================

# ===== Bot Definition Classes =====
class Intent:
    def __init__(self, name: str, parent: str = None) -> None:
        self.name = name
        self.parent_intent = parent
        self.sample_utterances = []
        self.slots = []

    def to_dict(self) -> dict:
        result = {'name': self.name}

        if self.parent_intent:
            result['parent'] = self.parent_intent
        if self.sample_utterances:
            result['sample_utterances'] = self.sample_utterances
        if self.slots:
            result['slots'] = [slot.to_dict() for slot in self.slots]

        return result


class BotDefinition:
    def __init__(self, name: str, locale: str) -> None:
        self.name = name
        self.locale = locale
        self.intent_clarification = []
        self.intents = []
        self.inherited_intents = []
        self.slot_types = []

    def get_intent_by_name(self, intent_name: str) -> Intent:
        result = None
        for intent in self.intents:
            if intent.name == intent_name:
                result = intent
                break
        return result

    def to_dict(self) -> Dict:
        result = {
            'name': self.name,
            'locale': self.locale,
            'intent_clarification': self.intent_clarification,
            'intents': [],
            'slot_types': [slot_type.to_dict() for slot_type in self.slot_types]
        }

        for intent in self.intents:
            result['intents'].append(intent.to_dict())
        for inherited_intent in self.inherited_intents:
            result['intents'].append(inherited_intent.to_dict())

        return result


class Slot:
    def __init__(self, name: str, type: str, prompt: str, required: bool) -> None:
        self.name = name
        self.type = type
        self.prompt = prompt
        self.required = required

    def to_dict(self) -> dict:
        return {
            'name': self.name,
            'type': self.type,
            'prompt': self.prompt,
            'required': self.required
        }


class SlotType:
    def __init__(self, name: str, values: List[str] = []) -> None:
        self.name = name
        self.values = values

    def to_dict(self) -> dict:
        return {
            'name': self.name,
            'values': self.values
        }


# ==================================

# ====== Bot Template Classes ======

class Script:
    def __init__(self, agent: str, sample_response: str, slot_to_elicit: str = None,
                 intent_to_elicit: str = None, confirm_intent: bool = False,
                 assume_intent: bool = False, close: bool = False) -> None:
        self.agent = agent
        self.sample_response = sample_response
        self.slot_to_elicit = slot_to_elicit
        self.intent_to_elicit = intent_to_elicit
        self.confirm_intent = confirm_intent
        self.assume_intent = assume_intent
        self.close = close

    def to_dict(self) -> Dict:
        result = {
            'agent': self.agent,
            'sample_response': False if not self.sample_response else self.sample_response,
            'slot_to_elicit': False if not self.slot_to_elicit else self.slot_to_elicit,
            'intent_to_elicit': False if not self.intent_to_elicit else self.intent_to_elicit,
            'confirm_intent': self.confirm_intent,
            'assume_intent': self.assume_intent,
            'close': self.close,
        }

        return result


class Conversation:
    def __init__(self, name: str, scenario_id: str, customer_instructions: str, description: str, bias: List[str]) -> None:
        self.name = name
        self.scenario_id = str(scenario_id)
        self.customer_instructions = customer_instructions
        self.description = description
        self.bias = bias
        self.scripts = []

    def to_dict(self) -> Dict:
        result = {
            'conversation': None,
            'name': self.name,
            'scenario_id': self.scenario_id,
            'bias': self.bias,
            'customer_instructions': self.customer_instructions,
            'description': self.description,
            'script': [script.to_dict() for script in self.scripts]
        }

        return result


class BotTemplate:
    def __init__(self, name: str, locale: str, domain: str) -> None:
        self.name = name
        self.locale = locale
        self.domain = domain
        self.general_agent_instructions = "In this task, you will be playing the Agent side of a customer service bot."
        self.general_customer_instructions = "In this task, you will be playing the Customer side of a customer service bot."
        self.slot_filled_instructions = "IMPORTANT - Do not re-ask for information that the Customer gave you when they first made the request. Just skip over that prompt when you get to it."
        self.custom_slot_instructions = "Please use the following information to answer the bot's questions."
        self.personal_information = "IMPORTANT - Do not give any personal information to the bot! If it asks you for personal information, just make up something that sounds realistic."
        self.agent_did_not_understand = "Sorry, I did not understand. Goodbye!"
        self.conversations = []

    def get_conversation_by_name(self, conversation_name: str) -> Conversation:
        result = None
        for conversation in self.conversations:
            if conversation.name == conversation_name:
                result = conversation
                break
        return result

    def to_dict(self) -> Dict:
        result = {
            'name': self.name,
            'locale': self.locale,
            'domain': self.domain,
            'general_agent_instructions': self.general_agent_instructions,
            'general_customer_instructions': self.general_customer_instructions,
            'slot_filled_instructions': self.slot_filled_instructions,
            'custom_slot_instructions': self.custom_slot_instructions,
            'personal_information': self.personal_information,
            'agent_did_not_understand': self.agent_did_not_understand,
            'conversations': [conversation.to_dict() for conversation in self.conversations]
        }

        return result

# ==================================

# ======== Common Functions ========


def get_header_indices(worksheet: Worksheet, header_values: list) -> Tuple:
    header_row = None
    header_cols = {}
    for row in worksheet:
        values = [col.value for col in row if not isinstance(col,  EmptyCell)]
        if values and all([val in header_values for val in values]):
            header_row = row[0].row
            for header_val in header_values:
                header_cols[header_val] = values.index(header_val) + 1
            break

    if not header_row:
        logger.info(f"{worksheet.title=}, {header_values=}, {header_cols=}")
        raise IncompleteWorksheet(worksheet)

    return header_row, header_cols

# ==================================

# ==== Bot Definition Functions ====


def parse_bot_def_sheet(worksheet: Worksheet) -> BotDefinition:
    bot_name = None
    bot_locale = None

    for row in worksheet.rows:
        if row[0].value == 'Bot Name':
            bot_name = row[1].value
            if not bot_name:
                raise MissingValue(worksheet, row[1])
        elif row[0].value == 'Language':
            bot_locale = row[1].value
            if not bot_locale:
                raise MissingValue(worksheet, row[1])

    if not bot_name or not bot_locale:
        raise IncompleteWorksheet(worksheet)

    return BotDefinition(bot_name, bot_locale)


def parse_intents_sheet(worksheet: Worksheet) -> List[Intent]:
    try:
        first_cell: Cell = [
            row[0] for row in worksheet.rows if row[0].value == 'Intent Name'][0]
    except IndexError:
        raise IncompleteWorksheet(worksheet)

    start_row = first_cell.row + 1
    col = first_cell.column
    intents = []
    existing_intent_names = set()
    for row in worksheet.iter_rows(min_row=start_row, min_col=col, max_col=col, values_only=True):
        if row[0] not in existing_intent_names:
            intents.append(Intent(row[0]))
            existing_intent_names.add(row[0])

    return intents


def parse_inherited_intents_sheet(worksheet: Worksheet) -> List[Intent]:
    try:
        first_cell: Cell = [row[0] for row in worksheet.rows if row[0].value ==
                            'Intent Name' and row[1].value == 'Parent Intent'][0]
    except IndexError:
        raise IncompleteWorksheet(worksheet)

    start_row = first_cell.row + 1
    col = first_cell.column
    intents = []
    for row in worksheet.iter_rows(min_row=start_row, min_col=col, max_col=col+1, values_only=True):
        intents.append(Intent(row[0], row[1]))

    return intents


def parse_sample_utterances_worksheet(worksheet: Worksheet) -> List[Tuple]:
    try:
        first_cell: Cell = [row[0] for row in worksheet.rows if row[0].value ==
                            'Intent Name' and row[1].value == 'Sample Utterance'][0]
    except IndexError:
        raise IncompleteWorksheet(worksheet)

    start_row = first_cell.row + 1
    col = first_cell.column
    sample_utterances = []
    for row in worksheet.iter_rows(min_row=start_row, min_col=col, max_col=col+1, values_only=True):
        sample_utterances.append((row[0], row[1]))

    return sample_utterances


def parse_slots_worksheet(worksheet: Worksheet) -> Dict[str, List[Slot]]:
    header_row, header_cols = get_header_indices(
        worksheet,
        ['Intent Name',	'Slot Name', 'Required', 'Type', 'Prompt'])

    slots = {}
    min_row = header_row + 1
    min_col = min(header_cols.values())
    max_col = max(header_cols.values())

    for row in worksheet.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        intent_name = row[header_cols['Intent Name'] - 1].value
        slot_name = row[header_cols['Slot Name'] - 1].value
        required = row[header_cols['Required'] - 1].value
        slot_type = row[header_cols['Type'] - 1].value
        prompt = row[header_cols['Prompt'] - 1].value

        if intent_name not in slots:
            slots[intent_name] = []

        slots[intent_name].append(Slot(
            name=slot_name,
            type=slot_type,
            prompt=prompt,
            required=required
        ))

    return slots


def parse_slot_types_worksheet(worksheet: Worksheet) -> List[SlotType]:
    logger.info(f"{worksheet.__dict__}")

    header_row, header_cols = get_header_indices(
        worksheet, ['Slot Type Name',	'Slot Value'])

    logger.info(f"{header_row=}, {header_cols=}")

    slot_type_dict = {}
    min_row = header_row + 1
    min_col = min(header_cols.values())
    max_col = max(header_cols.values())

    for row in worksheet.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        name = row[header_cols['Slot Type Name'] - 1].value
        value = row[header_cols['Slot Value'] - 1].value
        logger.info(f"{name=}, {value=}")
        slot_type_dict.setdefault(name, []).append(value)
        '''
        if name not in slot_type_dict:
            slot_type_dict[name] = []
        slot_type_dict[name].append(value)
        '''

    return [SlotType(name, values) for name, values in slot_type_dict.items()]


def parse_intent_clarification_worksheet(worksheet: Worksheet) -> List[str]:
    header_row, header_cols = get_header_indices(
        worksheet, ['Intent Clarification'])

    intent_clarifications = []
    min_row = header_row + 1
    min_col = min(header_cols.values())
    max_col = max(header_cols.values())

    for row in worksheet.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        intent_clarifications.append(row[header_cols['Intent Clarification'] - 1].value
        )

    return intent_clarifications


def bot_definition_to_yaml(infile_path, outfile_path):
    if not outfile_path:
        outfile_path = os.path.splitext(infile_path)[0] + '.yaml'

    wb = load_workbook(filename=infile_path, read_only=True)

    try:
        bot_definition = parse_bot_def_sheet(wb['Bot Definition'])
        bot_definition.intents = parse_intents_sheet(wb['Intents'])
        bot_definition.inherited_intents = parse_inherited_intents_sheet(
            wb['Inherited Intents'])
        sample_utterances = parse_sample_utterances_worksheet(
            wb['Sample Utterances'])

        intent_names = [intent.name for intent in bot_definition.intents]
        for utterance in sample_utterances:
            if utterance[0] not in intent_names:
                raise InvalidIntentName(utterance[0])

            intent = bot_definition.get_intent_by_name(utterance[0])
            intent.sample_utterances.append(utterance[1])

        slots = parse_slots_worksheet(wb['Slots'])
        for intent_name, slot_list in slots.items():
            intent = bot_definition.get_intent_by_name(intent_name)
            intent.slots = slot_list

        bot_definition.slot_types = parse_slot_types_worksheet(wb['Slot Types'])
        bot_definition.intent_clarification = parse_intent_clarification_worksheet(
            wb['Intent Clarification'])
    except KeyError as e:
        logger.info(f"KeyError: {e}")
        raise IncompleteWorkbook()

    wb.close()

    yaml = YAML()
    yaml.width = 200
    yaml.indent(mapping=2, sequence=4, offset=2)
    with open(outfile_path, mode='w', encoding='utf-8') as out_file:
        yaml.dump(bot_definition.to_dict(), out_file)

# ==================================

# ===== Bot Template Functions =====


def parse_bot_template_sheet(worksheet: Worksheet) -> BotTemplate:
    bot_name = None
    bot_locale = None
    bot_domain = None

    for row in worksheet.rows:
        if row[0].value == 'Template Name':
            bot_name = row[1].value
            if not bot_name:
                raise MissingValue(worksheet, row[1])
        elif row[0].value == 'Language':
            bot_locale = row[1].value
            if not bot_locale:
                raise MissingValue(worksheet, row[1])
        elif row[0].value == 'Domain':
            bot_domain = row[1].value
            if not bot_domain:
                raise MissingValue(worksheet, row[1])

    if not bot_name or not bot_locale or not bot_domain:
        raise IncompleteWorksheet(worksheet)

    return BotTemplate(bot_name, bot_locale, bot_domain)


def parse_conversations_worksheet(worksheet: Worksheet) -> List[Conversation]:
    header_row, header_cols = get_header_indices(
        worksheet,
        [
            'Conversation Name',
            'Conversation Type (SingleIntent, MultiIntent, etc.)',
            'Scenario ID',
            'Customer Instructions',
            'Description'
        ])

    conversations = []
    min_row = header_row + 1
    min_col = min(header_cols.values())
    max_col = max(header_cols.values())

    for row in worksheet.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        conversation_name = row[header_cols['Conversation Name'] - 1].value
        bias = row[header_cols['Conversation Type (SingleIntent, MultiIntent, etc.)'] - 1].value
        scenario_id = row[header_cols['Scenario ID'] - 1].value
        customer_instructions = row[header_cols['Customer Instructions'] - 1].value
        description = row[header_cols['Description'] - 1].value

        conversations.append(Conversation(
            name=conversation_name,
            scenario_id=scenario_id,
            customer_instructions=customer_instructions,
            description=description,
            bias=[bias]
        ))

    return conversations


def parse_scripts_worksheet(worksheet: Worksheet) -> Dict[str, List[Script]]:
    header_row, header_cols = get_header_indices(
        worksheet,
        [
            'Conversation Name',
            'Agent',
            'Sample Response',
            'Slot to Elicit',
            'Intent to Elicit',
            'Confirm Intent',
            'Assume Intent',
            'Close'
        ])

    scripts = {}
    min_row = header_row + 1
    min_col = min(header_cols.values())
    max_col = max(header_cols.values())

    for row in worksheet.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        if all([not cell.value for cell in row]):
            break

        conversation_name = row[header_cols['Conversation Name'] - 1].value
        agent = row[header_cols['Agent'] - 1].value
        sample_response = row[header_cols['Sample Response'] - 1].value
        slot_to_elicit = row[header_cols['Slot to Elicit'] - 1].value
        intent_to_elicit = row[header_cols['Intent to Elicit'] - 1].value
        confirm_intent = row[header_cols['Confirm Intent'] - 1].value
        assume_intent = row[header_cols['Assume Intent'] - 1].value
        close = row[header_cols['Close'] - 1].value

        if conversation_name not in scripts:
            scripts[conversation_name] = []

        scripts[conversation_name].append(Script(
            agent=agent,
            sample_response=sample_response,
            slot_to_elicit=slot_to_elicit,
            intent_to_elicit=intent_to_elicit,
            confirm_intent=confirm_intent,
            assume_intent=assume_intent,
            close=close
        ))

    return scripts


def bot_template_to_yaml(infile_path, outfile_path):
    if not outfile_path:
        outfile_path = os.path.splitext(infile_path)[0] + '.yaml'

    wb = load_workbook(filename=infile_path, read_only=True)

    try:
        bot_template = parse_bot_template_sheet(wb['Bot Template'])
        bot_template.conversations = parse_conversations_worksheet(
            wb['Conversations'])

        scripts = parse_scripts_worksheet(wb['Scripts'])
    except KeyError:
        raise IncompleteWorkbook()

    for conversation_name, script_list in scripts.items():
        conversation = bot_template.get_conversation_by_name(conversation_name)
        conversation.scripts = script_list

    wb.close()

    yaml=YAML()
    yaml.width = 200
    yaml.indent(mapping=2, sequence=4, offset=2)
    with open(outfile_path, mode='w', encoding='utf-8') as out_file:
        yaml.dump(bot_template.to_dict(), out_file)

# ==================================


def definition_command(args):
    if not os.path.isfile(args.infile) or not args.infile.endswith('.xlsx'):
        raise InvalidInputFile()

    bot_definition_to_yaml(args.infile, args.outfile)


def template_command(args):
    if not os.path.isfile(args.infile) or not args.infile.endswith('.xlsx'):
        raise InvalidInputFile()

    bot_template_to_yaml(args.infile, args.outfile)


def cli():
    parser = argparse.ArgumentParser(
        prog=__title__, description=__description__)

    parser.set_defaults(func=lambda _args: parser.print_help())
    parser.add_argument('--version', action='version',
                        version='%(prog)s v{}'.format(__version__))
    subparsers = parser.add_subparsers(help='Available sub-commands')

    definition_parser = subparsers.add_parser(
        'definition', help='Convert a Bot Definition Excel file to Yaml')
    definition_parser.add_argument(
        '--outfile', '-o', help="Path of generated Yaml file")
    definition_parser.add_argument('infile', help="Path of input Excel file")
    definition_parser.set_defaults(func=definition_command)

    template_parser = subparsers.add_parser(
        'template', help='Convert a Bot Template Excel file to Yaml')
    template_parser.add_argument(
        '--outfile', '-o', help="Path of generated Yaml file")
    template_parser.add_argument('infile', help="Path of input Excel file")
    template_parser.set_defaults(func=template_command)

    args = parser.parse_args()

    # Call sub-parser function
    if hasattr(args, 'func'):
        args.func(args)

class LexYamlGenGui:

    def __init__(self, parent):

        parent.title("Lex Yaml Gen")
        self.working_dir = USER_DOCS_DIR

        nbook = ttk.Notebook(parent)
        nbook.grid(column=0, row=0, sticky=(tk.N, tk.W, tk.E, tk.S))

        ttk.Label(parent, text=f'Version: {__version__}').grid(column=0, row=1, sticky=tk.W)

        def_frame = ttk.Frame(nbook, padding="3 3 12 12")
        def_frame.grid(column=0, row=0, sticky=(tk.N, tk.W, tk.E, tk.S))
        templ_frame = ttk.Frame(nbook, padding="3 3 12 12")
        templ_frame.grid(column=0, row=0, sticky=(tk.N, tk.W, tk.E, tk.S))
        nbook.add(def_frame, text='Bot Definition')
        nbook.add(templ_frame, text='Bot Template')

        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)

        self.bot_def_in_filepath = tk.StringVar()

        ttk.Label(def_frame, text="Convert a Bot Definition Excel file into YAML").grid(
            column=1, row=1, sticky=tk.N)
        ttk.Label(def_frame, textvariable=self.bot_def_in_filepath, borderwidth=2,
                  relief="groove", width=100).grid(column=1, row=2, sticky=(tk.W, tk.E))
        ttk.Button(def_frame, text="Open", command=self.select_bot_def_file).grid(
            column=2, row=2, sticky=tk.W)
        ttk.Button(def_frame, text="Generate Yaml", command=self.gen_bot_def_yaml).grid(
            column=2, row=3, sticky=tk.W)

        ttk.Separator(def_frame, orient=tk.HORIZONTAL).grid(column=1, row=4, sticky=(tk.W, tk.E))
        ttk.Label(def_frame, text="Generate sample Bot Definition Excel file").grid(
            column=1, row=5, sticky=tk.N)
        ttk.Button(def_frame, text="Generate Sample Excel", command=self.gen_bot_def_sample).grid(
            column=1, row=6, sticky=tk.W)

        self.bot_template_in_filepath = tk.StringVar()

        ttk.Label(templ_frame, text="Convert a Bot Template Excel file into YAML").grid(
            column=1, row=1, sticky=tk.N)
        ttk.Label(templ_frame, textvariable=self.bot_template_in_filepath, borderwidth=2,
                  relief="groove", width=100).grid(column=1, row=2, sticky=(tk.W, tk.E))
        ttk.Button(templ_frame, text="Open", command=self.select_bot_template_file).grid(
            column=2, row=2, sticky=tk.W)
        ttk.Button(templ_frame, text="Generate Yaml",
                   command=self.gen_bot_template_yaml).grid(column=2, row=3, sticky=tk.W)

        ttk.Separator(templ_frame, orient=tk.HORIZONTAL).grid(column=1, row=4, sticky=(tk.W, tk.E))
        ttk.Label(templ_frame, text="Generate sample Bot Template Excel file").grid(
            column=1, row=5, sticky=tk.N)
        ttk.Button(templ_frame, text="Generate Sample Excel", command=self.gen_bot_template_sample).grid(
            column=1, row=6, sticky=tk.W)

        for child in def_frame.winfo_children():
            child.grid_configure(padx=5, pady=5)

        for child in templ_frame.winfo_children():
            child.grid_configure(padx=5, pady=5)

    def select_bot_def_file(self, *args):
        files = [('Excel Files', '*.xlsx')]
        in_filepath = filedialog.askopenfilename(filetypes=files, initialdir=self.working_dir)
        if in_filepath:
            self.bot_def_in_filepath.set(in_filepath)
            self.working_dir = os.path.dirname(in_filepath)

    def gen_bot_def_yaml(self, *args):
        files = [('Yaml Files', '*.yaml')]
        out_filepath = filedialog.asksaveasfilename(
            filetypes=files, defaultextension=files, initialdir=self.working_dir)
        if out_filepath:
            try:
                bot_definition_to_yaml(
                    self.bot_def_in_filepath.get(), out_filepath)

                self.working_dir = os.path.dirname(out_filepath)
                messagebox.showinfo(
                    message=f'Bot Definition YAML file saved to: {out_filepath}')
            except Error as err:
                messagebox.showerror(title=type(err).__name__, message=str(err))

    def select_bot_template_file(self, *args):
        files = [('Excel Files', '*.xlsx')]
        in_filepath = filedialog.askopenfilename(filetypes=files, initialdir=self.working_dir)
        if in_filepath:
            self.bot_template_in_filepath.set(in_filepath)
            self.working_dir = os.path.dirname(in_filepath)

    def gen_bot_template_yaml(self, *args):
        files = [('Yaml Files', '*.yaml')]
        out_filepath = filedialog.asksaveasfilename(
            filetypes=files, defaultextension=files, initialdir=self.working_dir)
        if out_filepath:
            try:
                bot_template_to_yaml(
                    self.bot_template_in_filepath.get(), out_filepath)

                self.working_dir = os.path.dirname(out_filepath)
                messagebox.showinfo(
                    message=f'Bot Template YAML file saved to: {out_filepath}')
            except Error as err:
                messagebox.showerror(title=type(err).__name__, message=str(err))

    def gen_bot_def_sample(self, *args):
        files = [('Excel Files', '*.xlsx')]
        out_filepath = filedialog.asksaveasfilename(
            filetypes=files, defaultextension=files, initialdir=self.working_dir)

        if out_filepath:
            copyfile(os.path.join(BASE_DIR, 'sample_bot_definition.xlsx'), out_filepath)
            self.working_dir = os.path.dirname(out_filepath)
            messagebox.showinfo(
                message=f'Sample Bot Definition Excel file saved to: {out_filepath}')

    def gen_bot_template_sample(self, *args):
        files = [('Excel Files', '*.xlsx')]
        out_filepath = filedialog.asksaveasfilename(
            filetypes=files, defaultextension=files, initialdir=self.working_dir)

        if out_filepath:
            copyfile(os.path.join(BASE_DIR, 'sample_bot_template.xlsx'), out_filepath)
            self.working_dir = os.path.dirname(out_filepath)
            messagebox.showinfo(
                message=f'Sample Bot Template Excel file saved to: {out_filepath}')


if __name__ == "__main__":
    # cli()
    root = tk.Tk()
    LexYamlGenGui(root)
    root.mainloop()
