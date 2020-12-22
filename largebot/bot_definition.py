import pandas as pd
from openpyxl import load_workbook
import glob
import os
import argparse

def batch_list(list_to_slice, batch_size):
    return [list_to_slice[i:i+batch_size] for i in range(0, len(list_to_slice), batch_size)]

def _parse_args():
    """
    Parses user command line arguments.
    """

    parser = argparse.ArgumentParser(
        description="Script to populate sample_bot_definition.xlsx files"
    )
    parser.add_argument(dest='infiles', nargs='+', 
                        help="one or more xlsx files containing intents, like \"12142020_EN-US_Training-Data_Drop1_IntS.xlsx\"")
    parser.add_argument('--template', '-t', dest='template', required=True,
                        help='path of the bot definition template you want to populate')
    parser.add_argument('--batch-number', '-n', dest='start', required=False, default=1,
                        help='number at which to start enumerating batches')
    parser.add_argument('--batch-size', '-s', dest='size', required=False, default=None,
                        help='number at which to start enumerating batches')
    parser.add_argument('--outdir', '-o', dest='outdir', required=True,
                        help='path of the output folder with the populated templates')
    args = parser.parse_args()
    return args

if __name__ == "__main__":
    args = _parse_args()
    infiles = []
    for x in args.infiles:
        if glob.glob(x) == []:
            infiles.append(x)
        else:
            infiles += glob.glob(x)
    
    for xlsx in infiles:
        
        dirname, filename = os.path.split(xlsx)
        
        print(f"Getting data from file \"{xlsx}\"...")
        training_data = pd.read_excel(xlsx, sheet_name="Intent_Slot Creation")
        # check for missing values
        if training_data.isnull().values.any():
            print("Warning: the file contains misssing values")
        # trim colum to make script more robust
        training_data["Slot Name Required or optional"] = training_data["Slot Name Required or optional"].str.strip()
        # replace "required" or "optional" with True or False. Adding also spaces 
        training_data["Slot Name Required or optional"] = training_data["Slot Name Required or optional"].map({'Required': True, 'Optional': False})
        # rename columns so they match the template file for improved readability
        training_data = training_data.rename(columns={"IntentName": "Intent Name", "Slot Name Required or optional": "Required", "Slot Value": "Type", "Slot Prompt": "Prompt"})
        # remove unneded columns
        training_data = training_data.reindex(columns=["Intent Name", "Slot Name", "Required", "Type", "Prompt"])

        training_data["Prompt"] = training_data["Prompt"].apply(lambda x: x.replace('"', ''))

        builtins = {
            'NUMBER': 'AMAZON.Number',
            'FIRSTNAME': 'AMAZON.FirstName',
            'LASTNAME': 'AMAZON.LastName',
            'PHONENUMBER': 'AMAZON.PhoneNumber',
            'ALPHANUMERIC': 'AMAZON.AlphaNumeric',
            'EMAILADDRESS': 'AMAZON.EmailAddress',
            'POSTALADDRESS': 'AMAZON.PostalAddress',
            'DATEINTERVAL': 'AMAZON.DateInterval',
            'TIME': 'AMAZON.Time',
            'DATE': 'AMAZON.Date',
            'CURRENCY': 'AMAZON.Currency',
            'DURATION': 'AMAZON.Duration',
            'STATE': 'AMAZON.State',
            'COUNTRY': 'AMAZON.Country',
            'CITY': 'AMAZON.City',
            'AIRPORT': 'AMAZON.Airport',
            'DAYOFWEEK': 'AMAZON.DayOfWeek',
            'PERCENTAGE': 'AMAZON.Percentage',
            'SPEED': 'AMAZON.Speed',
            'STREETNAME': 'AMAZON.StreetName',
            'WEIGHT': 'AMAZON.Weight'
        }

        training_data["Type"] = training_data["Type"].apply(
            lambda x: builtins.get(x, None) or x
        )
        # get unique list of intents
        intents = list(training_data["Intent Name"].unique())
        batch_size = int(args.size) if args.size else len(intents)
        print(f"{batch_size=}")
        batched_intents = batch_list(intents, batch_size)

        file_counter = int(args.start)
        for batch in batched_intents:
            # print((batch))
            df_batch = pd.DataFrame(batch)
            df = pd.DataFrame()
            for intent in batch:
                # print((intent))
                temp_df = training_data[training_data["Intent Name"] == intent]
                df = df.append(temp_df)

            # load bot_definition to populate
            book = load_workbook(args.template)
            # book = load_workbook('sample_bot_definition.xlsx')
            outfilename = os.path.splitext(filename)[0] + "_" + str(file_counter).zfill(3) + ".xlsx"
            # outname = "sample_bot_definition_" + str(file_counter).zfill(3) + ".xlsx"
            outpath = os.path.join(args.outdir, outfilename)
            writer = pd.ExcelWriter(outpath, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            for sheetname in writer.sheets:
                if sheetname == "Intents":
                    # print(f"Populating \"{sheetname}\" sheet...")
                    df_batch.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,header=False)
                elif sheetname == "Sample Utterances":
                    # print(f"Populating \"{sheetname}\" sheet...")
                    df_batch.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,header=False)
                elif sheetname == "Slots":
                    # print(f"Populating \"{sheetname}\" sheet...")
                    df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False,header=False)
                    writer.save()
                    print(f"Writing file \"{outpath}\"...")
                    file_counter += 1
        print(f"Done with file {xlsx}\n")
    print("Done!")

# script assumes tabnames and column names are the same and files are well-formed, i.e. with no missing values
# need to trim the whole dataframe